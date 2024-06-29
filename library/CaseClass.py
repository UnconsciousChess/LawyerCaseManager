import os,sys

# 将当前工作目录添加到系统路径中
sys.path.append(os.getcwd())

# 不要生成字节码
sys.dont_write_bytecode = True

# 导入nanoid模块
from nanoid import generate

# 导入诉讼参与人类
from .LitigantClass import *

class Case():

    def __init__(self) -> None:
        # 初始化案件类的各项属性，均设为初始值
        # 案件类型
        self._CaseType = 0
        # 诉讼标的额
        self._LitigationAmount = 0
        # 案由
        self._CaseOfAction = ""
        # 管辖法院
        self._JurisdictionDict = {
            "一审": "",
            "二审": "",
            "再审": "",
            "执行": "",
            "仲裁": "",
        }        
        # 上传文件列表
        self._UploadFilesList = {}
        # 诉讼请求（上诉请求）
        self._ClaimText = ""
        # 事实与理由
        self._FactAndReasonText = ""
        # 案件文件所在文件夹路径
        self._CaseFolderPath = ""
        # 原告主体列表,如原告[0]（原告一）、原告[1]（原告二）...
        self._PlaintiffList = []
        # 被告主体列表,如被告[0]（被告一）、被告[1]（被告二）...
        self._DefendantList = []
        # 第三人主体列表,如第三人[0]（第三人一）、第三人[1]（第三人二）...
        self._LegalThirdPartyList = []
        # 调解意向
        self._MediationIntention = False
        # 拒绝调解理由
        self._RejectMediationReasonText = ""
        # 案件代理的阶段
        self._CaseAgentStage = []
        # 风险代理情况
        self._RiskAgentStatus = None
        # 风险代理前期费用
        self._RiskAgentUpfrontFee = 0
        # 风险代理后期比例
        self._RiskAgentPostFeeRate = 0
        # 非风险代理的固定费用(是一个列表，第一个元素对应第一个阶段的费用，第二个元素对应第二个阶段的费用...)
        self._AgentFixedFeeList = []
        # 案号
        self._CaseCourtCode = ""
        # 案件id（实例化案件时自动生成，用于前后端交互时识别）,该属性没有对应的Set方法，也无法input
        self._CaseId = "Case-" + generate(alphabet='0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz', size=16)


        # 下面通过读取文件得到一些通用列表,并作为类的属性

        # 准备所有合法法院名称的列表，方便进行对比，以防止输入的法院名称有误
        with open(r"Data\PublicInfomationList\CourtNameList.txt","r",encoding="utf-8") as f:
            StandardJurisdictionList = f.readlines()
            # 去除每个法院名称的换行符
            self._StandardJurisdictionList = [i.strip() for i in StandardJurisdictionList]
        # 准备所有合法民事案由的列表，方便进行对比，以防止输入的案由有误
        with open(r"Data\PublicInfomationList\CauseOfActionList.txt","r",encoding="utf-8") as f:
            CaseOfActionList = f.readlines()
            # 去除每个案由的换行符，并放到列表CauseOf中
            self._CaseOfActionList = [i.strip() for i in CaseOfActionList]


    # 下面是案件类的外部方法,包括:
    # a.获取各属性的方法（Get-xxxx)
    # b.设定各属性的方法(Set-xxxx、Append-xxx)
    # c.输入案件信息的方法(Input-xxx)
    # d.输出案件信息的方法(Output-xxx)

    # ==================Get方法：下面定义外部直接获取各属性的方法====================

    # 案件类型
    def GetCaseType(self):
        return self._CaseType
    # 诉讼标的额
    def GetLitigationAmount(self):
        return self._LitigationAmount
    # 案由
    def GetCauseOfAction(self):
        return self._CaseOfAction
    # 管辖法院
    def GetJurisdictionDict(self):
        return self._JurisdictionDict
    # 上传文件列表
    def GetUploadFilesList(self):
        return self._UploadFilesList
    # 诉讼请求
    def GetClaimText(self):
        return self._ClaimText
    # 事实与理由
    def GetFactAndReasonText(self):
        return self._FactAndReasonText
    # 案件文件所在文件夹路径
    def GetCaseFolderPath(self):
        return self._CaseFolderPath
    # 原告主体列表
    def GetPlaintiffList(self):
        return self._PlaintiffList
    # 被告主体列表
    def GetDefendantList(self):
        return self._DefendantList
    # 第三人主体列表
    def GetLegalThirdPartyList(self):
        return self._LegalThirdPartyList
    # 调解意愿
    def GetMediationIntention(self):
        return self._MediationIntention
    # 拒绝调解理由
    def GetRejectMediationReasonText(self):
        return self._RejectMediationReasonText
    # 案件代理的阶段
    def GetCaseAgentStage(self):
        return self._CaseAgentStage
    # 风险代理情况
    def GetRiskAgentStatus(self):
        return self._RiskAgentStatus
    # 风险代理前期费用
    def GetRiskAgentUpfrontFee(self):
        return self._RiskAgentUpfrontFee
    # 风险代理后期比例
    def GetRiskAgentPostFeeRate(self):
        return self._RiskAgentPostFeeRate
    # 非风险代理的固定费用数组
    def GetAgentFixedFee(self):
        return self._AgentFixedFeeList
    # 法院案号
    def GetCaseCourtCode(self):
        return self._CaseCourtCode
    # 案件id
    def GetCaseId(self):
        return self._CaseId
    

    # ===== Get方法：下面定义外部获取各属性的一些进阶方法（主要涉及输出一些常用字符串）=====

    # 返回所有原告的名称字符串的函数
    def GetAllPlaintiffNames(self) -> str:
        Plaintiffs = ""
        for litigant in self._PlaintiffList:
            Plaintiffs += litigant.GetName() + "、"
        Plaintiffs = Plaintiffs[:-1]
        return Plaintiffs
    
    # 返回所有被告的名称字符串的函数
    def GetAllDefendantNames(self) -> str:
        Defendants = ""
        for litigant in self._DefendantList:
            Defendants += litigant.GetName() + "、"
        Defendants = Defendants[:-1]
        return Defendants
    
    # 返回代理阶段代码对应的中文字符串
    def GetCaseAgentStageStr(self) -> str:
        CaseAgentStageOutputString = ""
        if self._CaseAgentStage == []:
            return CaseAgentStageOutputString
        for i in self._CaseAgentStage:
            if i == 1:
                CaseAgentStageOutputString += "一审立案阶段、"
            elif i == 2:
                CaseAgentStageOutputString += "一审开庭阶段、"
            elif i == 3:
                CaseAgentStageOutputString += "二审阶段、"
            elif i == 4:
                CaseAgentStageOutputString += "执行阶段、"
            elif i == 5:
                CaseAgentStageOutputString += "再审阶段、"
        # 去掉最后一个顿号
        CaseAgentStageOutputString = CaseAgentStageOutputString[:-1]
        return CaseAgentStageOutputString        
    
    # 返回我方当事人列表及代理方向
    def GetOurClientListAndSide(self) -> list:
        OurClientList = []
        OurClientSide = ""
        for plaintiff in self.GetPlaintiffList():
            if plaintiff.IsOurClient():
                OurClientList.append(plaintiff)
                OurClientSide = "p"
        # 如果我方不是代理原告，则为代理被告
        if OurClientSide != "p" :
            for defendant in self.GetDefendantList():
                if defendant.IsOurClient():
                    OurClientList.append(defendant)
                    OurClientSide = "d"

        return OurClientList,OurClientSide

    def GetOurClientNames(self) -> str:
        OurClientList,OurClientSide = self.GetOurClientListAndSide()
        OurClientNames = ""
        for litigant in OurClientList:
            OurClientNames += litigant.GetName() + "、"
        OurClientNames = OurClientNames[:-1]
        return OurClientNames
    
    def GetCourtNameStr(self) -> str:
        CourtNameStr = ""
        for stage,court in self._JurisdictionDict.items():
            # 如果法院不为空，则添加到字符串中
            if court != "":
                CourtNameStr += stage + "：" + court + "、"
        CourtNameStr = CourtNameStr[:-1]
        return CourtNameStr

    # ============Set和Append方法：下面定义设定各属性的方法（含输入值校验）=============

    # 案件类型设定方法，1为民事案件，2为行政案件，3为执行案件
    def SetCaseType(self,CaseType,debugmode=False):
        if isinstance(CaseType,int):
            if CaseType == 1 or CaseType == 2 or CaseType == 3:
                self._CaseType = CaseType
            else:
                if debugmode:
                    print("SetCaseType报错：参数只能为1 2 3 ")
        else:
            if debugmode:
                print("SetCaseType报错：参数必须为整型") 
    
    # 诉讼标的额设定方法
    def SetLitigationAmount(self,LitigationAmount,debugmode=False):
        # 尝试将输入值转换为浮点数
        try:
            LitigationAmount = float(LitigationAmount)
        except:
            if debugmode:
                print("SetLitigationAmount报错：输入值并非浮点数")
            return
        # 诉讼标的额不能小于零
        if LitigationAmount >= 0:
            self._LitigationAmount = LitigationAmount
        else:
            if debugmode:
                print("SetLitigationAmount报错：诉讼标的不能小于零")
            return

    # 案由设定方法
    def SetCauseOfAction(self,CaseOfAction,debugmode=False):

        if (CaseOfAction in self._CaseOfActionList):
            self._CaseOfAction = CaseOfAction
            if debugmode:
                print("输入的案由【%s】添加成功" % CaseOfAction)
        else:
            if debugmode:
                print(" SetCauseOfActionb报错：输入的【%s】名称不符合现有民事、行政、执行案由规定,请重新输入" % CaseOfAction)
    
    # 管辖法院设定方法
    def SetJurisdictionDict(self,InputJurisdictionDict,debugmode=False):

        for stage,jurisdictionname in InputJurisdictionDict.items():
             #如果输入的键名（阶段）在现有的字典的键名（阶段）中,则进一步判断输入的键值是否合法  
            if stage in self._JurisdictionDict.keys():      
                    # 判断输入的键名（法院名称）是否在Standard法院列表中，即是否合法  
                    if (jurisdictionname in self._StandardJurisdictionList):
                        self._JurisdictionDict[stage] = jurisdictionname
                        if debugmode:
                            print("添加【%s】管辖法院【%s】成功" % (stage,jurisdictionname))
                    else:
                        if debugmode:
                            print("SetJurisdictionDict报错：输入的案由【%s】不符合现有法院名称。" % jurisdictionname)
            else:
                if debugmode:
                    print("SetJurisdictionDict报错：输入的【%s】键名不符合规范" % stage)

    # 诉讼请求设定方法
    def SetClaimText(self,ClaimText,debugmode=False):
        if isinstance(ClaimText,str):
            self._ClaimText = ClaimText
        else:
            if debugmode:
                print("SetClaimText报错：该输入对象的类型与属性不匹配,诉讼请求输入值为字符串")

    # 事实与理由设定方法
    def SetFactAndReasonText(self,FactAndReasonText,debugmode=False):
        if isinstance(FactAndReasonText,str):
            self._FactAndReasonText = FactAndReasonText
        else:
            if debugmode:
                print("SetFactAndReasonText报错：该输入对象的类型与属性不匹配,事实与理由输入值为字符串")

    # 案件案件生成文件夹路径设定方法
    def SetCaseFolderPath(self,CaseFolderPath,debugmode=False):
        if isinstance(CaseFolderPath,str):
            if os.path.exists(CaseFolderPath):
                # 判断输入的路径是否为文件路径
                if os.path.isfile(CaseFolderPath) == False:
                    self._CaseFolderPath = CaseFolderPath
                    if debugmode:
                        print("SetCaseFolderPath报错：案件文件所在文件夹路径设定成功")
                else:
                    if debugmode:
                        print("SetCaseFolderPath报错：输入的路径为文件路径，请重新输入")
            else:
                if debugmode:
                    print("SetCaseFolderPath报错：文件夹不存在，请重新输入")
        else:
            if debugmode:
                print("SetCaseFolderPath报错：该输入对象的类型与属性不匹配,案件文件所在文件夹路径输入值为字符串")
    
    # 备注设定方法
    def SetCommentText(self,Comment,debugmode=False):
        if isinstance(Comment,str):
            self._Comment = Comment
        else:
            if debugmode:
                print("SetCommentText报错：该输入对象的类型与属性不匹配,备注输入值为字符串")

    # 调解意愿设定方法
    def SetMediationIntention(self,MediationIntention,debugmode=False):
        if isinstance(MediationIntention,bool):
            self._MediationIntention = MediationIntention
        elif isinstance(MediationIntention,str):
            if MediationIntention == "True" or MediationIntention == "true" or MediationIntention == "TRUE" or MediationIntention == "1":
                MediationIntention = True
            elif MediationIntention == "False" or MediationIntention == "false" or MediationIntention == "FALSE" or MediationIntention == "0":
                MediationIntention = False
            self._MediationIntention = MediationIntention
        else:
            if debugmode:
                print("SetMediationIntention报错：该输入对象的类型与属性不匹配,调解意愿输入值为布尔值或字符串")

    # 拒绝理由设定方法
    def SetRejectMediationReasonText(self,RejectReason,DebugMode=False):
        if isinstance(RejectReason,str):
            self._RejectReason = RejectReason
        else:
            if DebugMode:
                print("SetRejectMediationReasonText报错：该输入对象的类型与属性不匹配,拒绝理由输入值为字符串")

    # 案件代理阶段设定方法
    def SetCaseAgentStage(self,CaseAgentStage,DebugMode=False):
        if isinstance(CaseAgentStage,list):
            # 对比新列表的数字是否与现有的列表重复
            for i in CaseAgentStage:
                if i not in self._CaseAgentStage and i in [1,2,3,4,5]:
                    self._CaseAgentStage.append(i)
        elif isinstance(CaseAgentStage,int):
            if CaseAgentStage not in self._CaseAgentStage and CaseAgentStage in [1,2,3,4,5]:
                self._CaseAgentStage.append(CaseAgentStage)
        else:
            if DebugMode:
                print("SetCaseAgentStage报错：该输入对象的类型与属性不匹配,案件代理阶段输入值为列表或1-5的整数")

    # 风险代理情况设定方法
    def SetRiskAgentStatus(self,RiskAgentStatus,DebugMode=False):
        if isinstance(RiskAgentStatus,bool):
            self._RiskAgentStatus = RiskAgentStatus
        elif isinstance(RiskAgentStatus,str):
            if RiskAgentStatus == "True" or RiskAgentStatus == "true" or RiskAgentStatus == "TRUE" or RiskAgentStatus == "1":
                self._RiskAgentStatus = True
            elif RiskAgentStatus == "False" or RiskAgentStatus == "false" or RiskAgentStatus == "FALSE" or RiskAgentStatus == "0":
                self._RiskAgentStatus = False
        else:
            if DebugMode:
                print("SetRiskAgentStatus报错：该输入对象的类型与属性不匹配,风险代理情况输入值为布尔值或字符串")
    
    # 风险代理前期费用设定方法
    def SetRiskAgentUpfrontFee(self,RiskAgentUpfrontFee,DebugMode=False):
        # 尝试将输入值转换为浮点数
        try:
            RiskAgentUpfrontFee = float(RiskAgentUpfrontFee)
        except:
            if DebugMode:
                print("输入值并非浮点数")
            return
        # 诉讼标的额不能小于零
        if RiskAgentUpfrontFee >= 0:
            self._RiskAgentUpfrontFee = RiskAgentUpfrontFee
        else:
            if DebugMode:
                print("风险代理前期费用不能小于零")
            return
    
    # 风险代理后期比例设定方法
    def SetRiskAgentPostFeeRate(self,RiskAgentPostFeeRate,DebugMode=False):
        # 尝试将输入值转换为浮点数
        try:
            RiskAgentPostFeeRate = float(RiskAgentPostFeeRate)
        except:
            if DebugMode:
                print("SetRiskAgentPostFeeRate报错：输入值并非浮点数")
            return
        # 诉讼标的额不能小于零
        if RiskAgentPostFeeRate >= 0:
            self._RiskAgentPostFeeRate = RiskAgentPostFeeRate
        else:
            if DebugMode:
                print("SetRiskAgentPostFeeRate报错：风险代理后期比例不能小于零")
            return

    # 非风险代理的固定费用设定方法（方法一直接输入列表）
    def SetAgentFixedFeeByList(self,AgentFixedFeeList,DebugMode=False):
        # 判断是否原本列表式否非空
        if self._AgentFixedFeeList == []:
            # 判断输入值是否为列表
            if isinstance(AgentFixedFeeList,list):
                # 再次判断列表的数量是否小于等于代理阶段的数量
                if len(AgentFixedFeeList) <= len(self.GetCaseAgentStage()):
                    self._AgentFixedFeeList = AgentFixedFeeList
                else:
                    if DebugMode:
                        print("SetAgentFixedFee报错：输入的固定费用列表数量超出已有的代理阶段数量")
            else:
                if DebugMode:
                    print("SetAgentFixedFee报错：该输入对象的类型与属性不匹配,非风险代理的固定费用输入值为列表或整数")
        else:
            if DebugMode:
                print("SetAgentFixedFee报错：非风险代理的固定费用列表已经存在，无法再次设定")
    
    # 非风险代理的固定费用设定方法（方法二直接输入代理费用和代理阶段）
    def SetAgentFixedFeeByAdd(self,AgentFixedFee,Stage,DebugMode=False):
        # 判断输入值是否为整数
        if isinstance(AgentFixedFee,float) and isinstance(Stage,int):
            if Stage <= len(self.GetCaseAgentStage()-1):
                self._AgentFixedFeeList[Stage] = AgentFixedFee
            else:
                if DebugMode:
                    print("SetAgentFixedFee报错：输入的阶段序号超出已有的阶段序号")
        else:
            if DebugMode:
                print("SetAgentFixedFee报错：输入对象的类型与属性不匹配,非风险代理的固定费用输入值为列表或整数")
        
    # 法院案号设定方法
    def SetCaseCourtCode(self,CaseCourtCode):
        if isinstance(CaseCourtCode,str):
            self._CaseCourtCode = CaseCourtCode
        else:
            print("SetCaseCourtCode方法报错：该输入对象的类型与属性不匹配,法院案号输入值为字符串")


     # 将诉讼参与人添加到案件中
    

    # 添加诉讼参与人的方法
    def AppendLitigant(self,ALitigant,PrintLogMode=False) -> bool:
        # 先判断添加进来的东西是什么
        if isinstance(ALitigant,Litigant):
            # 判断诉讼地位是原告、被告还是第三人
            # 原告1  被告2  第三人3
            if ALitigant.GetLitigantPosition() == 1:
                self._PlaintiffList.append(ALitigant)
                if PrintLogMode:
                    print("添加原告【%s】成功" % ALitigant.GetName())   
                return True
            elif ALitigant.GetLitigantPosition()  == 2:
                self._DefendantList.append(ALitigant)
                if PrintLogMode:
                    print("添加被告【%s】成功" % ALitigant.GetName())
                return True
            elif ALitigant.GetLitigantPosition() == 3:
                self._LegalThirdPartyList.append(ALitigant)
                if PrintLogMode:
                    print("添加第三人【%s】成功" % ALitigant.GetName()) 
                return True
            else:
                if PrintLogMode:
                    print("当前添加的诉讼参与人【%s】缺失诉讼地位参数LitigantPosition，无法添加到列表之中" % ALitigant.GetName())
                return False
   
    # ===========Input方法：下面定义批量输入案件信息的方法=============

    # 设定一个类的内部方法，对于参数键名和键值，分别进行处理（键值为中文，来源于txt和excel文件）
    def SetCaseInfoWithKeyAndValue(self,Key,Value):
        # 根据Key的不同，调用不同的设定方法
        if Key == '案件类型':
            self.SetCaseType(int(Value))

        elif Key == '诉讼标的额':
            self.SetLitigationAmount(float(Value))

        elif Key == '案由':
            self.SetCauseOfAction(str(Value))

        elif Key == '管辖法院':
            # 将Value转换为字典，以便SetJurisdictionDict方法处理

            # 以逗号分割字符串,分割出Stage:Court的字符串列表
            StageAndCourtList = Value.split(",")
            for stageandcourt in StageAndCourtList: 
                # 去掉阶段和法院名称的空格
                stageandcourt = stageandcourt.strip()
                # 去掉“阶段”两个字
                stageandcourt = stageandcourt.replace("阶段","")
                # 以冒号分割字符串,中文冒号或英文冒号
                if "：" in stageandcourt:
                    Stage,Court = stageandcourt.split("：")
                elif ":" in stageandcourt:
                    Stage,Court = stageandcourt.split(":")
                # 调用设定管辖法院的方法
                self.SetJurisdictionDict({Stage:Court})

        elif Key == '诉讼请求':
            self.SetClaimText(str(Value))

        elif Key == '事实与理由':
            self.SetFactAndReasonText(str(Value))

        elif Key == '案件文件所在文件夹路径':
            self.SetCaseFolderPath(Value)

        elif Key == '调解意愿':
            self.SetMediationIntention(Value)
            
        elif Key == '拒绝调解理由':
            self.SetRejectMediationReasonText(str(Value))

        elif Key == '案件代理阶段':
            # 将Value转换为数字，以便SetCaseAgentStage方法处理
            # 以逗号分割字符串,分割出各阶段的字符串列表
            StageList = Value.split(",")
            for stagestr in StageList:
                if "一审" in stagestr:
                    self.SetCaseAgentStage([1,2])

                elif "二审" in stagestr:
                    self.SetCaseAgentStage(3)
                elif "执行" in stagestr:
                    self.SetCaseAgentStage(4)
                elif "再审" in stagestr:
                    self.SetCaseAgentStage(5)

        elif Key == '风险代理情况':
            self.SetRiskAgentStatus(Value)

        elif Key == '风险代理前期费用':
            self.SetRiskAgentUpfrontFee(float(Value))

        elif Key == '风险代理后期比例':
            self.SetRiskAgentPostFeeRate(float(Value))

        elif Key == '非风险代理的固定费用':
            # 将Value转换为列表
            # 以逗号分割字符串,分割出各阶段的字符串列表
            FixedFeeList = Value.split(",")
            # 将字符串列表转换为浮点数列表
            FixedFeeList = [float(i) for i in FixedFeeList]
            self.SetAgentFixedFeeByList(FixedFeeList)
            
        elif Key == '法院案号':
            self.SetCaseCourtCode(Value)

    # 设定一个类的内部方法，针对从前端读回来的信息（字典），逐个键值对进行处理，本方法后期重点维护  
    def InputCaseInfoFromWebDict(self,Key,Value):

        # 案件类型
        if Key == 'caseType':
            self.SetCaseType(Value)

        # 诉讼标的额
        elif Key == 'litigationAmount':
            self.SetLitigationAmount(Value)

        # 案由
        elif Key == 'caurseOfAction':
            self.SetCauseOfAction(Value)

        # 管辖法院
        elif Key == 'courtName':
            # 以逗号分割字符串,分割出Stage:Court的字符串列表（这是暂时的，假设输入的是一个字符串）
            # 后期要在前端做好多选框，这里的Value应该是一个字典或者字典
            StageAndCourtList = Value.split(",")
            for stageandcourt in StageAndCourtList: 
                # 去掉阶段和法院名称的空格
                stageandcourt = stageandcourt.strip()
                # 去掉“阶段”两个字
                stageandcourt = stageandcourt.replace("阶段","")
                # 以冒号分割字符串,中文冒号或英文冒号
                if "：" in stageandcourt:
                    Stage,Court = stageandcourt.split("：")
                elif ":" in stageandcourt:
                    Stage,Court = stageandcourt.split(":")
                # 调用设定管辖法院的方法
                self.SetJurisdictionDict({Stage:Court})


        # 诉讼请求
        elif Key == 'claimText':
            self.SetClaimText(Value)

        # 事实与理由
        elif Key == 'factAndReasonText':
            self.SetFactAndReasonText(Value)

        # 案件文件所在文件夹路径
        elif Key == 'caseFolderGeneratedPath':
            self.SetCaseFolderPath(Value)

        # 调解意愿
        elif Key == 'mediationIntention':
            self.SetMediationIntention(Value)

        # 拒绝调解理由            
        elif Key == 'rejectMediationReasonText':
            self.SetRejectMediationReasonText(Value)

        # 原告信息（目前是测试只有一个人的时候）
        elif Key == 'plaintiffInfoPath':
            plaintiff = Litigant()
            plaintiff.InputLitigantInfoFromTxt(Value)
            
            # 将该参与人设置为原告
            plaintiff.SetLitigantPosition(1)
            # 测试用，将原告选定为我方当事人(应当删除)

            plaintiff.SetOurClient(True)
            # 再次读取原告信息并添加到原告列表中 
            plaintiff.InputLitigantInfoFromTxt(Value)
            self.AppendLitigant(plaintiff)
        
        # 被告信息（目前是测试只有一个人的时候）
        elif Key == 'defendantInfoPath':
            defendant = Litigant()
            defendant.InputLitigantInfoFromTxt(Value)
            # 将该参与人设置为被告
            defendant.SetLitigantPosition(2)
            # 测试用，将被告选定为对方当事人（应当删除）
            defendant.SetOurClient(False)
            # 再次读取被告信息并添加到被告列表中 
            defendant.InputLitigantInfoFromTxt(Value)
            self.AppendLitigant(defendant)

        # 案件代理的阶段
        elif Key == 'caseAgentStage':
            # 进来的value就是一个列表，因此可以直接调用SetCaseAgentStage方法
            self.SetCaseAgentStage(Value)

        # 风险代理情况
        elif Key == 'riskAgentStatus':
            self.SetRiskAgentStatus(Value)

        # 风险代理前期费用
        elif Key == 'riskAgentUpfrontFee':
            self.SetRiskAgentUpfrontFee(Value)

        # 风险代理后期比例
        elif Key == 'riskAgentPostFeeRate':
            self.SetRiskAgentPostFeeRate(Value)

        # 非风险代理的固定费用
        elif Key == 'agentFixedFee':
            pass
            # self.SetAgentFixedFeeByList(FixedFeeList)

        # 法院案号
        elif Key == 'caseCourtCode':
            self.SetCaseCourtCode(Value)       



    # 读取一个txt文档的路径来输入上述案件信息
    def InputCaseInfoFromTxt(self,CaseInfoFilePath,DebugMode=False):
        # 判断路径是否存在
        if not os.path.exists(CaseInfoFilePath):
            print("文件路径不存在")
            return
        
        # 路径存在，则读取文件
        with open(CaseInfoFilePath,"r",encoding="utf-8") as f:
            CaseInfoLines = f.readlines()
        # 去除每个案由的换行符
        CaseInfoLines = [i.strip() for i in CaseInfoLines]
        # 逐行读取文件
        for line in CaseInfoLines:
            # 判断是否为空行，是则跳过
            if line == "":
                continue
            # 判断是否为以#开头的注释行，是则跳过
            if line[0] == "#":
                continue
            # 判断是否有等于号，没有则跳过
            if "=" not in line:
                continue
            # 以等于号分割字符串
            Key,Value = line.split("=")
            # 对上述分割出来的键值对进行处理
            self.SetCaseInfoWithKeyAndValue(Key,Value)

    # 读取一个excel文档的路径来输入上述案件信息
    def InputCaseInfoFromExcel(self,CaseInfoFilePath,DebugMode=False):
        from openpyxl import load_workbook
        wb = load_workbook(CaseInfoFilePath)
        ws = wb.active
        # 读取第一列的数据
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=1,values_only=True):
            for cell in row:
                # 判断是否为空行，是则跳过
                if cell == "":
                    continue
                # 第一列的数据是键名
                Key = cell.value
                # 第二列的数据是键值
                Value = ws.cell(row=cell.row,column=2).value
                # 对上述分割出来的键值对进行处理
                self.SetCaseInfoWithKeyAndValue(Key,Value)

    # 从应用的前端中输入案件信息,重点维护
    def InputCaseInfoFromFrontEnd(self,CaseInfoDict,DebugMode=False):
        # 判断传入的参数是否为字典
        if isinstance(CaseInfoDict,dict):
            # 对于字典中的逐个键值对读取，并调用SetCaseInfoWithKeyAndValue方法对键值对进行处理
            for Key,Value in CaseInfoDict.items():
                self.SetCaseInfoFromWebDict(Key,Value)


    # ===========Output方法：下面定义输出案件信息的方法=============

    # 输出案件信息到txt文件
    def OutputCaseInfoToTxt(self,DebugMode=False):
        # 输出文件路径=案件文件夹路径中获取
        OutputFilePath = self.GetCaseFolderPath() 
        # 判断路径是否存在
        if not os.path.exists(OutputFilePath):
            if DebugMode:
                print("文件路径不存在")

        # 判断路径是否以\结尾,如果不是则加上\
        if not OutputFilePath.endswith("\\"):
            OutputFilePath += "\\output.txt"
        else:
            OutputFilePath += "output.txt"

        # 写入文件
        with open(file=OutputFilePath,mode="w",encoding="utf-8") as f:
            # 逐个输出案件信息
            if self.GetCaseType() == 1:
                f.write("案件类型=民事案件\n")
            elif self.GetCaseType() == 2:
                f.write("案件类型=行政案件\n")
            elif self.GetCaseType() == 3:
                f.write("案件类型=执行案件\n")
            f.write("诉讼标的=%s元\n" % self.GetLitigationAmount())
            f.write("案由=%s\n" % self.GetCauseOfAction())
            f.write("管辖法院=")
            for stage,court in self.GetJurisdictionDict().items():
                f.write("%s:%s," % (stage,court))
            f.write("\n")
            f.write("诉讼请求=%s\n" % self.GetClaimText())
            f.write("事实与理由=%s\n" % self.GetFactAndReasonText())
            f.write("案件文件所在文件夹路径=%s\n" % self.GetCaseFolderPath())
            f.write("原告主体列表=")
            for plaintiff in self.GetPlaintiffList():
                f.write("%s," % plaintiff.GetName())
            f.write("\n")
            f.write("被告主体列表=")
            for defendant in self.GetDefendantList():
                f.write("%s," % defendant.GetName())
            f.write("\n")
            f.write("第三人主体列表=")
            for thirdparty in self.GetLegalThirdPartyList():
                f.write("%s," % thirdparty.GetName())
            f.write("\n")
            f.write("调解意愿=%s\n" % self.GetMediationIntention())
            f.write("拒绝调解理由=%s\n" % self.GetRejectMediationReasonText())
            f.write("案件代理的阶段:%s\n" % self.GetCaseAgentStageStr())
            if self.GetRiskAgentStatus() == True:
                f.write("本案为风险代理。\n")
                f.write("风险代理前期费用=%s元\n" % self.GetRiskAgentUpfrontFee())
                f.write("风险代理后期比例=%s%%\n" % self.GetRiskAgentPostFeeRate())
            else:
                f.write("非风险代理的固定费用=")
                if self.GetAgentFixedFee() == []:
                    f.write("无\n")
                else:
                    for fee in self.GetAgentFixedFee():
                        f.write("%s," % fee)
            f.write("法院案号是：%s\n" % self.GetCaseCourtCode())
    
    # 输出案件信息到excel文件
    def OutputCaseInfoToExcel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font,Alignment


        # 输出文件路径=案件文件夹路径中获取
        OutputFilePath = self.GetCaseFolderPath()
        # 判断路径是否存在
        if not os.path.exists(OutputFilePath):
            return
        # 判断路径是否以\结尾,如果不是则加上\
        if not OutputFilePath.endswith("\\"):
            OutputFilePath += "\\"
        wb = Workbook()
        ws = wb.active
        # 表头
        ws.append(["案件具体属性","对应值"])
        # 调整表头格式，方便看
        for cell in ws[1]:
            # 将表头的字体加粗
            cell.font = Font(bold=True)
            # 居中
            cell.alignment = Alignment(horizontal='center',vertical='center')
        # 逐行输出案件信息
        # 根据案件类型的数字，输出对应的字符串

        if self.GetCaseType() == 1:
            ws.append(["案件类型","民事案件"])
        elif self.GetCaseType() == 2:
            ws.append(["案件类型","行政案件"])
        elif self.GetCaseType() == 3:
            ws.append(["案件类型","执行案件"])

        ws.append(["诉讼标的",self.GetLitigationAmount()])
        ws.append(["案由",self.GetCauseOfAction()])
        # 管辖法院变成字符串
        JurisdictionStr = ""
        for stage,court in self.GetJurisdictionDict().items():
            JurisdictionStr += stage + ":" + court + ","
        ws.append(["管辖法院",JurisdictionStr])
        ws.append(["诉讼请求",self.GetClaimText()])
        ws.append(["事实与理由",self.GetFactAndReasonText()])
        ws.append(["案件文件所在文件夹路径",self.GetCaseFolderPath()])
        # 先判断原告主体列表是否为空
        if self.GetPlaintiffList() == []:
            ws.append(["原告主体列表","无"])
        else:
            ws.append(["原告主体列表",self.GetAllPlaintiffNames()])
        # 先判断被告主体列表是否为空
        if self.GetDefendantList() == []:
            ws.append(["被告主体列表","无"])
        else:
            ws.append(["被告主体列表",self.GetAllDefendantNames()])
        # 先判断第三人主体列表是否为空
        if self.GetLegalThirdPartyList() == []:
            ws.append(["第三人主体列表","无"])
        else:
            ws.append(["第三人主体列表",self.GetLegalThirdPartyList()])

        # 根据调解意愿的布尔值，输出对应的字符串
        if self.GetMediationIntention() == True:
            ws.append(["调解意愿","愿意调解"])
        else:
            ws.append(["调解意愿","拒绝调解"])

        ws.append(["拒绝调解理由",self.GetRejectMediationReasonText()])
        ws.append(["案件代理的阶段",self.GetCaseAgentStageStr()])
        if self.GetRiskAgentStatus() == True:
            ws.append(["风险代理前期费用",self.GetRiskAgentUpfrontFee()])
            ws.append(["风险代理后期比例",self.GetRiskAgentPostFeeRate()])
        else:
            ws.append(["非风险代理的固定费用",self.GetAgentFixedFee()])
        ws.append(["法院案号",self.GetCaseCourtCode()])

        # 设定文件名
        # OutputName = self.GetAllPlaintiffNames() + "诉" + self.GetAllDefendantNames() + "案件信息.xlsx"
        OutputName = "案件信息.xlsx"
        # 保存文件
        wb.save(OutputFilePath + OutputName)
        print("案件信息已经保存到%s" % (OutputFilePath + OutputName))


    # 输出案件信息到前端（直接输出字典）
    def OutputCaseInfoToFrontEnd(self,DebugMode=False) -> dict:
        # 需要返回的字典初始化
        OutputDict = {}

        # 逐个输出案件信息
        OutputDict["caseCourtCode"] = self.GetCaseType()
        OutputDict["litigationAmount"] = self.GetLitigationAmount()
        OutputDict["causeOfAction"] = self.GetCauseOfAction()
        OutputDict["courtName"] = self.GetCourtNameStr()    #法院名称直接输出字符串
        OutputDict["claimText"] = self.GetClaimText()
        OutputDict["factAndReason"] = self.GetFactAndReasonText()
        OutputDict["caseFolderGeneratedPath"] = self.GetCaseFolderPath()
        OutputDict["mediationIntention"] = self.GetMediationIntention()
        OutputDict["rejectMediationReasonText"] = self.GetRejectMediationReasonText()
        OutputDict["caseAgentStage"] = self.GetCaseAgentStage()
        OutputDict["caseType"] = self.GetCaseType()
        OutputDict["caseCourtCode"] = self.GetCaseCourtCode()
        OutputDict["caseId"] = self.GetCaseId()

        # 根据是否为风险收费代理，输出不同的费用信息
        if self.GetRiskAgentStatus() == True:     #风险收费
            OutputDict["riskAgentStatus"] = self.GetRiskAgentStatus()
            OutputDict["riskAgentUpfrontFee"] = self.GetRiskAgentUpfrontFee()
            OutputDict["riskAgentPostFeeRate"] = self.GetRiskAgentPostFeeRate()
            OutputDict["agentFixedFee"] = ""
        else:
            OutputDict["riskAgentStatus"] = self.GetRiskAgentStatus()
            OutputDict["riskAgentUpfrontFee"] = ""
            OutputDict["riskAgentPostFeeRate"] = ""
            OutputDict["agentFixedFee"] = self.GetAgentFixedFee()

        # 测试
        print(OutputDict)
        # 返回字典
        return OutputDict
    

    # 输出当前案件的当事人信息到前端
    def OutputLitigantInfoToFrontEnd(self,DebugMode=False) -> dict:

        # 需要返回的字典初始化
        OutputDict = {}

        # 原告主体列表（列表归零）
        LitigantList = []
        for plaintiff in self.GetPlaintiffList():
            LitigantList.append(plaintiff.OutputLitigantInfoToFrontEnd())
        OutputDict["plaintiffs"] = LitigantList 
        
        # 被告主体列表（列表重新归零）
        LitigantList = []
        for defendant in self.GetDefendantList():
            LitigantList.append(defendant.OutputLitigantInfoToFrontEnd())
        OutputDict["defendants"] = LitigantList

        # 第三人主体列表（列表重新归零）
        LitigantList = []
        for thirdparty in self.GetLegalThirdPartyList():
            LitigantList.append(thirdparty.OutputLitigantInfoToFrontEnd())
        OutputDict["thirdParties"] = LitigantList

        # 返回字典
        return OutputDict